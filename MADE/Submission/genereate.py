import openpyxl

workbook= openpyxl.load_workbook('Movie Cataloge Indo.xlsx')
movies = workbook['Movie']

last_row = movies.max_row
data=[]
data.append('<array name=\"movies_drawable\">\n')
data.append('<string-array name=\"movies_title\">\n')
data.append('<string-array name=\"movies_description\">\n')
data.append('<string-array name=\"movies_score\">\n')
data.append('<string-array name=\"movies_status\">\n')
data.append('<string-array name=\"movies_release_information\">\n')
data.append('<string-array name=\"movies_language\">\n')
data.append('<string-array name=\"movies_runtime\">\n')
data.append('<string-array name=\"movies_budget\">\n')
data.append('<string-array name=\"movies_revenue\">\n')
data.append('<string-array name=\"movies_genre\">\n')
for row in range(2, last_row):
    #poster
    data[0] += '\t<item>@drawable/movie_' + movies['A' + str(row)].value + "</item>\n"
    data[1] += '\t<item>' + movies['B' + str(row)].value + "</item>\n"
    data[2] += '\t<item>' + movies['C' + str(row)].value + "</item>\n"
    data[3] += '\t<item>' + str(movies['D' + str(row)].value) + "</item>\n"
    data[4] += '\t<item>' + movies['E' + str(row)].value + "</item>\n"
    data[5] += '\t<item>' + movies['F' + str(row)].value + "</item>\n"
    data[6] += '\t<item>' + movies['G' + str(row)].value + "</item>\n"
    data[7] += '\t<item>' + movies['H' + str(row)].value + "</item>\n"
    data[8] += '\t<item>' + str(movies['I' + str(row)].value) + "</item>\n"
    data[9] += '\t<item>' + str(movies['J' + str(row)].value) + "</item>\n"
    data[10] += '\t<item>' + movies['K' + str(row)].value + "</item>\n"

data[0] += "</array>"
for index in range(1, len(data)):
    data[index] += "</string-array>"
movies_file = open("movies.txt", 'a')
movies_file.write(
    data[0] + "\n" +
    data[1] + "\n" +
    data[2] + "\n" +
    data[3] + "\n" +
    data[4] + "\n" +
    data[5] + "\n" +
    data[6] + "\n" +
    data[7] + "\n" +
    data[8] + "\n" +
    data[9] + "\n" +
    data[10] + "\n")
movies_file.close
tvshows = workbook['TV Series']

last_row = tvshows.max_row
data=[]
data.append('<array name=\"tvshows_drawable\">\n')
data.append('<string-array name=\"tvshows_title\">\n')
data.append('<string-array name=\"tvshows_description\">\n')
data.append('<string-array name=\"tvshows_score\">\n')
data.append('<string-array name=\"tvshows_status\">\n')
data.append('<string-array name=\"tvshows_language\">\n')
data.append('<string-array name=\"tvshows_runtime\">\n')
data.append('<string-array name=\"tvshows_genre\">\n')
for row in range(2, last_row):
    #poster
    data[0] += '\t<item>@drawable/tvshow_' + tvshows['A' + str(row)].value + "</item>\n"
    data[1] += '\t<item>' + tvshows['B' + str(row)].value + "</item>\n"
    data[2] += '\t<item>' + tvshows['C' + str(row)].value + "</item>\n"
    data[3] += '\t<item>' + str(tvshows['D' + str(row)].value) + "</item>\n"
    data[4] += '\t<item>' + tvshows['E' + str(row)].value + "</item>\n"
    data[5] += '\t<item>' + tvshows['F' + str(row)].value + "</item>\n"
    data[6] += '\t<item>' + tvshows['G' + str(row)].value + "</item>\n"
    data[7] += '\t<item>' + tvshows['H' + str(row)].value + "</item>\n"

data[0] += "</array>"
for index in range(1, len(data)):
    data[index] += "</string-array>"
tvshows_file = open("tvshows.txt", 'a')
tvshows_file.write(
    data[0] + "\n" +
    data[1] + "\n" +
    data[2] + "\n" +
    data[3] + "\n" +
    data[4] + "\n" +
    data[5] + "\n" +
    data[6] + "\n" +
    data[7] + "\n" )
tvshows_file.close
